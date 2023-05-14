import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter.filedialog import askopenfilename
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


def getDayOfWeek(date):
    date_str = date.replace('.', '/')
    return datetime.strptime(date_str, '%d/%m/%Y').strftime('%A')


def isDate(str):
    try:
        date_str = str.replace('.', '/')
        datetime.strptime(date_str, '%d/%m/%Y')
        return True
    except:
        return False


def getFileName():
    filename.set(askopenfilename())


# Hard coded column names
# Original file has info in brackets on a separate row
date_col = 'ДАТА'
hour = 'ЧАС'
el_energy = 'КОЛИЧЕСТВО ЕЛ. ЕНЕРГИЯ (кВтч)'
price_bneb = 'ЦЕНА БНЕБ (лв./кВтч)'
administrative_tax_percentage = 'ДОГОВОРЕНА АДМИНИСТРАТИВНА ТАКСА (%)'
min_admin_tax = 'МИНИМАЛНА АДМИНИСТРАТИВНА ТАКСА (ЛВ./кВтч.)'
admin_tax = 'АДМИНИСТРАТИВНА ТАКСА (ЛВ./кВтч.)'
energy_price = 'ОБЩА ЦЕНА ЗА ЕЛ. ЕНЕРГИЯ (лв./кВтч)'
owed_price_energy = 'ДЪЛЖИМА СУМА ЗА ЕЛ. ЕНЕРГИЯ (лв.)'
owed_price_minus_tax = 'Дължима сума без такси (лв.)'


# Creates the GUI
root = Tk()
form = tk.Frame(root, padx=100, pady=100)
form.grid()
filename = tk.StringVar()
tk.Label(form, text="Select the file that you want graphs from").grid(
    column=0, row=0)
tk.Button(form, text="Select file", command=getFileName).grid(column=0, row=1)
root.mainloop()
# while True:
#   root.update()


# hardcoded file location -> to be replaced with the filename received
# from selecting the file through the GUI
file_location = 'C:/Users/DimitarParpulov/Downloads/energoPro_may-clean.xlsx'
file_location_copy = filename.get()
# data = pd.read_excel(file_location_copy)
# df = pd.DataFrame(data)
# data_grouped_by_date = df.groupby(date_col).agg(lambda x: list(x))
# days_in_month = len(df[date_col].unique())

# need to make the df and get days_in_month


def createCleanedDataExcelFile(input_file, workbook):
    original_worksheet = load_workbook(input_file).worksheets[0]
    worksheet = workbook.worksheets[0]
    worksheet.title = 'ЕнергоПро'

    col_names = (
        date_col,
        hour,
        el_energy,
        price_bneb,
        administrative_tax_percentage,
        min_admin_tax,
        admin_tax,
        energy_price,
        owed_price_energy,
        owed_price_minus_tax
    )
    worksheet.append(col_names)

    starting_row = 11

    selected_row = list(original_worksheet.rows)[11]

    # Clean and add data to first sheet
    # Adding the missing 11 rows because it starts from different than first row in the sheet
    for i in range(starting_row, days_in_month * 24 + starting_row):
        selected_row = list(original_worksheet.rows)[i]
        if isDate(selected_row[1].value):
            row_data = (
                selected_row[1].value,
                int(selected_row[2].value),
                selected_row[3].value,
                selected_row[4].value,
                selected_row[5].value,
                selected_row[6].value,
                selected_row[7].value,
                selected_row[8].value,
                selected_row[9].value,
                selected_row[9].value -
                (selected_row[3].value * selected_row[4].value)
            )
            worksheet.append(row_data)
        else:
            break


def createBarChart(workbook, start_row, end_row, graph_start_cell):
    sheet_charts = workbook['Почасово потребление']
    sheet_data = workbook.worksheets[0]
    data = Reference(sheet_data, min_col=3, max_col=3,
                     min_row=start_row, max_row=end_row)
    categories = Reference(sheet_data, min_col=2,
                           min_row=start_row, max_row=end_row)
    date = list(sheet_data.rows)[start_row][0].value

    chart = BarChart()
    chart.height = 10
    chart.width = 18
    chart.add_data(data)
    # chart.add_data(data, titles_from_data=True)
    chart.legend = None
    chart.set_categories(categories)
    # chart.shape = 4
    chart.type = "col"
    # chart.style = 10
    chart.title = date + ' ' + getDayOfWeek(date)
    chart.x_axis.title = 'Hour of day'
    chart.y_axis.title = 'Energy consumption'
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 240
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    sheet_charts.add_chart(chart, graph_start_cell)


starting_cells = [
    "A1", "L1", "W1", "AH1", "AS1", "BD1", "BO1",
    "A21", "L21", "W21", "AH21", "AS21", "BD21", "BO21",
    "A41", "L41", "W41", "AH41", "AS41", "BD41", "BO41",
    "A61", "L61", "W61", "AH61", "AS61", "BD61", "BO61",
    "A81", "L81", "W81", "AH81", "AS81", "BD81", "BO81"
]

new_workbook = Workbook()
new_workbook.create_sheet('Месечна справка')
new_workbook.create_sheet('Почасово потребление')

min_row = 2
max_row = 25
step = 24

createCleanedDataExcelFile(file_location_copy, new_workbook)
for i in range(days_in_month):
    try:
        createBarChart(new_workbook, min_row, max_row, starting_cells[i])
    except:
        break
    min_row += step
    max_row += step

new_workbook.save('results.xlsx')


# print(df.loc[:, el_energy])


# Overview from second sheet on main screen
# Show graph for specific date
