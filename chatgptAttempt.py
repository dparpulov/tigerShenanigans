import sys
from datetime import datetime, date
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill, Font
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QProgressBar
from decimal import Decimal

# Shared column names
date_col = 'ДАТА'
el_energy = 'КОЛИЧЕСТВО ЕЛ. ЕНЕРГИЯ (кВтч)'
owed_price_energy = 'ДЪЛЖИМА СУМА ЗА ЕЛ. ЕНЕРГИЯ (лв.)'

# Input data column names
hour = 'ЧАС'
price_bneb = 'ЦЕНА БНЕБ (лв./кВтч)'
administrative_tax_percentage = 'ДОГОВОРЕНА АДМИНИСТРАТИВНА ТАКСА (%)'
min_admin_tax = 'МИНИМАЛНА АДМИНИСТРАТИВНА ТАКСА (ЛВ./кВтч.)'
admin_tax = 'АДМИНИСТРАТИВНА ТАКСА (ЛВ./кВтч.)'
energy_price = 'ОБЩА ЦЕНА ЗА ЕЛ. ЕНЕРГИЯ (лв./кВтч)'
owed_price_minus_tax = 'Дължима сума без такси (лв.)'

# Output data column names
amount_el_energy_daytime = 'КОЛИЧЕСТВО ЕЛ. ЕНЕРГИЯ (м/у 08:00-16:00) (кВтч)'
owed_price_energy_daytime = 'ДЪЛЖИМА СУМА ЗА ЕЛ. ЕНЕРГИЯ (м/у 08:00-16:00) (лв.)'


def getDayOfWeek(date_string):
    return datetime.strptime(date_string, '%d.%m.%Y').strftime('%A')


def is_valid_date(date_string):
    try:
        datetime.strptime(date_string, '%d.%m.%Y')
        return True
    except:
        return False


def count_unique_dates(data):
    dates = set()
    for row in data:
        dates.add(row[0])

    return len(dates)


def check_is_day_off(date_str) -> bool:
    parsed_date = datetime.strptime(date_str, '%d.%m.%Y')
    year = parsed_date.year
    month = parsed_date.month
    day = parsed_date.day

    date_value = datetime(year, month, day)
    is_weekend = date_value.weekday() in (5, 6)

    holidays_22 = [
        date(year, 1, 3),
        date(year, 3, 3),
        date(year, 4, 22),
        date(year, 4, 25),
        date(year, 5, 2),
        date(year, 5, 6),
        date(year, 5, 24),
        date(year, 9, 6),
        date(year, 9, 22),
        date(year, 12, 26),
        date(year, 12, 27),
        date(year, 12, 28)
    ]
    holidays_23 = [
        date(year, 1, 2),
        date(year, 3, 3),
        date(year, 4, 14),
        date(year, 4, 17),
        date(year, 5, 1),
        date(year, 5, 8),
        date(year, 5, 24),
        date(year, 9, 6),
        date(year, 9, 22),
        date(year, 12, 25),
        date(year, 12, 26),
        date(year, 12, 27)
    ]
    holidays_24 = [
        date(year, 1, 1),
        date(year, 3, 4),
        date(year, 5, 1),
        date(year, 5, 3),
        date(year, 5, 6),
        date(year, 5, 24),
        date(year, 9, 6),
        date(year, 9, 23),
        date(year, 12, 24),
        date(year, 12, 25),
        date(year, 12, 26)
    ]
    if year == 2022:
        return date(year, month, day) in holidays_22 or is_weekend
    if year == 2023:
        return date(year, month, day) in holidays_23 or is_weekend
    if year == 2024:
        return date(year, month, day) in holidays_24 or is_weekend


def read_excel_data(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(min_row=11):
        # Get the value of the first cell in the row (column 2)
        date_string = row[1].value

        if is_valid_date(date_string):
            # If the value is a valid date, append the row data to the data list
            data.append([cell.value for cell in row[1:10]])

    return data


def create_excel_file(data):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "ЕнергоПро"
    # Create the second and third sheets
    sheet2 = wb.create_sheet('Месечна справка')
    wb.create_sheet('Почасово потребление')

    col_names_first_sheet = (
        date_col,
        hour,
        el_energy,
        price_bneb,
        administrative_tax_percentage,
        min_admin_tax,
        admin_tax,
        energy_price,
        owed_price_energy,
        owed_price_minus_tax
    )
    sheet1.append(col_names_first_sheet)
    col_names_second_sheet = (
        date_col,
        el_energy,
        owed_price_energy,
        amount_el_energy_daytime,
        owed_price_energy_daytime
    )
    sheet2.append(col_names_second_sheet)

    # Add the data to the first sheet
    for i, row in enumerate(data, start=2):
        for j, cell in enumerate(row, start=1):
            if j == 1:
                sheet1.cell(row=i, column=j, value=cell)
            if j == 2:
                sheet1.cell(row=i, column=j, value=int(cell))
            if j > 2:
                sheet1.cell(row=i, column=j, value=Decimal(cell))

        value = sheet1.cell(row=i, column=9).value - sheet1.cell(row=i,
                                                                 column=3).value * sheet1.cell(row=i, column=4).value

        sheet1.cell(row=i, column=10, value=Decimal(value))

    return wb


def createBarChart(excel_file, start_row, end_row, graph_start_cell):
    sheet_data = excel_file.worksheets[0]
    sheet_charts = excel_file.worksheets[2]
    data = Reference(sheet_data, min_col=3, max_col=3,
                     min_row=start_row, max_row=end_row)
    categories = Reference(sheet_data, min_col=2,
                           min_row=start_row, max_row=end_row)
    date = sheet_data.cell(row=start_row, column=1).value

    chart = BarChart()
    chart.height = 10
    chart.width = 18
    chart.add_data(data)
    # chart.add_data(data, titles_from_data=True)
    chart.legend = None
    chart.set_categories(categories)
    # # chart.shape = 4
    chart.type = "col"
    # # chart.style = 10
    chart.title = date + ' ' + getDayOfWeek(date)
    chart.x_axis.title = 'Hour of day'
    chart.y_axis.title = 'Energy consumption'
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 240
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    sheet_charts.add_chart(chart, graph_start_cell)


def calculateMonthlyResults(excel_file, start_row, step, days):
    sheet_data = excel_file.worksheets[0]
    sheet_results = excel_file.worksheets[1]
    el_enegry_amount = 0
    owed_money_amount = 0
    el_energy_amount_daytime = 0
    owed_money_amount_daytime = 0
    el_enegry_amount_total = 0
    owed_money_amount_total = 0
    el_energy_amount_daytime_total = 0
    owed_money_amount_daytime_total = 0
    el_enegry_amount_total_day_off = 0
    owed_money_amount_total_day_off = 0
    el_energy_amount_daytime_total_day_off = 0
    owed_money_amount_daytime_total_day_off = 0

    hour_counter = start_row
    hour_counter_daytime = start_row + 7

    for i in range(days):
        date = sheet_data.cell(row=start_row, column=1).value
        hour_counter = start_row
        hour_counter_daytime = start_row + 7
        is_day_off = check_is_day_off(date)

        for y in range(step):
            el_enegry_amount += sheet_data.cell(
                row=hour_counter, column=3).value
            el_enegry_amount_total += sheet_data.cell(
                row=hour_counter, column=3).value
            owed_money_amount += sheet_data.cell(
                row=hour_counter, column=9).value
            owed_money_amount_total += sheet_data.cell(
                row=hour_counter, column=9).value
            if is_day_off:
                el_enegry_amount_total_day_off += sheet_data.cell(
                    row=hour_counter, column=3).value
                owed_money_amount_total_day_off += sheet_data.cell(
                    row=hour_counter, column=9).value

            hour_counter += 1

        for k in range(9):
            el_energy_amount_daytime += sheet_data.cell(
                row=hour_counter_daytime, column=3).value
            el_energy_amount_daytime_total += sheet_data.cell(
                row=hour_counter_daytime, column=3).value
            owed_money_amount_daytime += sheet_data.cell(
                row=hour_counter_daytime, column=9).value
            owed_money_amount_daytime_total += sheet_data.cell(
                row=hour_counter_daytime, column=9).value
            if is_day_off:
                el_energy_amount_daytime_total_day_off += sheet_data.cell(
                    row=hour_counter_daytime, column=3).value
                owed_money_amount_daytime_total_day_off += sheet_data.cell(
                    row=hour_counter_daytime, column=9).value

            hour_counter_daytime += 1

        data = (
            date,
            Decimal(el_enegry_amount).quantize(Decimal('0.01')),
            Decimal(owed_money_amount).quantize(Decimal('0.01')),
            Decimal(el_energy_amount_daytime).quantize(Decimal('0.01')),
            Decimal(owed_money_amount_daytime).quantize(Decimal('0.01'))
        )
        sheet_results.append(data)

        if is_day_off:
            for cell in sheet_results[i+2]:
                cell.fill = PatternFill(fgColor='92D050',
                                        fill_type='solid')
        el_enegry_amount = 0
        owed_money_amount = 0
        el_energy_amount_daytime = 0
        owed_money_amount_daytime = 0
        start_row += step

    data_total = (
        "Общо:",
        el_enegry_amount_total,
        owed_money_amount_total,
        el_energy_amount_daytime_total,
        owed_money_amount_daytime_total
    )
    sheet_results.append(data_total)
    for cell in sheet_results[sheet_results.max_row]:
        cell.font = Font(bold=True)

    data_total = (
        "Работни дни:",
        el_enegry_amount_total - el_enegry_amount_total_day_off,
        owed_money_amount_total - owed_money_amount_total_day_off,
        el_energy_amount_daytime_total - el_energy_amount_daytime_total_day_off,
        owed_money_amount_daytime_total - owed_money_amount_daytime_total_day_off
    )
    sheet_results.append(data_total)
    for cell in sheet_results[sheet_results.max_row]:
        cell.font = Font(bold=True)

    data_total = (
        "Почивни дни:",
        el_enegry_amount_total_day_off,
        owed_money_amount_total_day_off,
        el_energy_amount_daytime_total_day_off,
        owed_money_amount_daytime_total_day_off
    )
    sheet_results.append(data_total)
    for cell in sheet_results[sheet_results.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fgColor='92D050',
                                        fill_type='solid')


starting_cells = [
    "A1", "L1", "W1", "AH1", "AS1", "BD1", "BO1",
    "A21", "L21", "W21", "AH21", "AS21", "BD21", "BO21",
    "A41", "L41", "W41", "AH41", "AS41", "BD41", "BO41",
    "A61", "L61", "W61", "AH61", "AS61", "BD61", "BO61",
    "A81", "L81", "W81", "AH81", "AS81", "BD81", "BO81"
]


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initButton(self):
        self.button = QPushButton('Select Excel file', self)
        self.button.clicked.connect(self.show_file_selection_dialog)
        self.button.move(150, 150)

    def initProgressBar(self):
        self.progressBar = QProgressBar(self)
        self.progressBar.setGeometry(50, 50, 350, 25)
        self.setWindowTitle('QProgressBar')

    def initUI(self):
        self.setGeometry(50, 50, 500, 500)
        self.initButton()
        self.initProgressBar()
        self.show()

    def updateProgressBar(self, value):
        self.progressBar.setValue(value)

    def show_file_selection_dialog(self):
        min_row = 2
        max_row = 25
        step = 24
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Select Excel file", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            data = read_excel_data(file_name)
            output_file = create_excel_file(data)
            days_in_month = count_unique_dates(data)
            calculateMonthlyResults(output_file, min_row, step, days_in_month)
            for i in range(days_in_month):
                try:
                    createBarChart(output_file, min_row,
                                   max_row, starting_cells[i])
                except:
                    break
                min_row += step
                max_row += step
                main_window.updateProgressBar(
                    int((i + 1) / days_in_month * 100))

            output_file.save("output.xlsx")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
